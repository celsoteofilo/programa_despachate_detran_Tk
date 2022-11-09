# _____________ Bibiliotecas para FRONT___________________________________________
import pathlib
import tkinter
from tkinter import *  # importa janelas
from tkinter import ttk, messagebox  # importa o cromobox -> rolagem de dados

# _______________________Biblioteca para BANCO DE DADOS___________________________
from tkinter.ttk import Combobox

import openpyxl, xlrd
from openpyxl import Workbook
import pathy

# ______________________BIBLIOTECA  para DATA e HORA para o sistema _______________
import datetime as dt

# ___________________BIBLIOTECA PARA FUNCOES INTERNAS _____________________________


import funcao_chamar_pag_consulta_veiculo

# ____________________LISTAS GERADAS PRINCIPAIS PARA CONSUMO  DO SISTEMA ___________

lista_tipos = ["CARRO", "CARRO ELETRICO", "MOTO", "MOTO ELETRICA ", "CAMINHAO "]
lista_codigos = []
lista_produto_padrao = []
lista_cep = []
banco_de_dados = []

# _____________________JANELA PRINCIPAL ._____________________________________________

janela = Tk()
janela.geometry("800x800+1000+300")  # tamnanho da dela + centralizaçao da tela
janela.title('DESPACHANTE DIGITAL')
janela.resizable = False  # resizable=True movintar a tela e false trava tela
janela.confirm_close = True  # confirm_close=True, pergunta de confirmaçao fechar janela

#janela.configure(bg="#326273") #-->Fundo da tela (cor)

# exemplo de formato de texto usando (Label )
# label = janela, texto = "OLA TESTE ", font 'arial 13', bg = "#3262673" ,fg '#fff' --> formato,tipo,cor.place(x=20 , y=20)
# desta forma acima e mais faciel mexer com fonte e cor .
                # ________________________________CRIANDO A IMAGEM ________

bg = PhotoImage(file="v.png")
label1 = Label(janela, image=bg)
label1.place(x=250, y=500)
#------------------------------------------------------------



                #________________________FUNCAO BUSCA CEP  (API) ________

def cep_busca():
    import requests

    cep = entrada_cep.get()
    cep = cep.replace("-", "").replace(".", "").replace(" ", "")

    if len(cep) == 8:
        link = f'https://viacep.com.br/ws/{cep}/json/'
        requisicao = requests.get(link)
        print(requisicao)

        resposta = requisicao.json()

        uf = resposta['uf']
        cidade = resposta['localidade']
        bairro = resposta['bairro']
        logradouro = resposta['logradouro']
        complemento = resposta['complemento']

        print(resposta)
    else:
        messagebox.showinfo('<CEP>, INVALIDO!!!')

    texto_cep = f'''
   
    CEP INFORMADO : {cep}
    RUA: {logradouro}
    BAIRRO : {bairro}
    CIDADE :  {cidade}
    COMPLEMENTO {complemento}
    UF : {uf}
   
    '''
    texto_cadastro["text"] = texto_cep



            # _______________________FUNCAO A BUSCAR DO DETRAN (CONSTRUCAO AINDA )________
''' 
def inserir_codigo():
    nome = entrada.get()
    veiculo = botao_combobox.get()
    placa = placa_veiculo.get()


def cep_busca():
    import requests

    cep = entrada_cep.get()
    cep = cep.replace("-", "").replace(".", "").replace(" ", "")

    if len(cep) == 8:
        link = f'https://viacep.com.br/ws/{cep}/json/'
        requisicao = requests.get(link)
        print(requisicao)

        resposta = requisicao.json()

        uf = resposta['uf']
        cidade = resposta['localidade']
        bairro = resposta['bairro']
        logradouro = resposta['logradouro']
        complemento = resposta['complemento']

        print(resposta)
    else:
        messagebox.showinfo('<CEP>, INVALIDO!!!')

    texto_cep = f
   
    CEP INFORMADO : {cep}
    RUA: {logradouro}
    BAIRRO : {bairro}
    CIDADE :  {cidade}
    COMPLEMENTO {complemento}
    UF : {uf}
   
    texto_cadastro["text"] = texto_cep
'''
            # ___________________FUNCAO GERAR CODIGO DE REGISTRO________


def inserir_codigo():
    nome = entrada.get()
    veiculo = botao_combobox.get()
    placa = placa_veiculo.get()
    chassi = entrada_chassi.get()
    cep = entrada_cep.get()

    # __________________puxando data e hora_____________
    data_criacao = dt.datetime.now()
    data_criacao.strftime('%d/%m/%y--%H:%M')

    #____________________ criando um codigo_____________
    codigo = len(lista_codigos) + 1  # lendo o codigo + 1
    codigo_str = "COD-{}".format(codigo)  # formatando o codigo
    lista_codigos.append((codigo_str, nome, placa, cep, veiculo, placa, data_criacao))  # montando a lista em seguencia
    lista_produto_padrao.append((placa, data_criacao))

    texto = f'''
            
    Pessoa Cadastrado : {nome}
    Tipo :  {veiculo}
    Placa Veiculo : {placa}
    Numero do Chassi: {chassi}
    codigo gerado: {codigo} 

    '''
    texto_cadastro["text"] = texto

    chassi = entrada_chassi.get()
    cep = entrada_cep.get()

    # _______________________puxando data e hora________________
    data_criacao = dt.datetime.now()
    data_criacao.strftime('%d/%m/%y--%H:%M')

    # ________________________criando um codigo___________________
    codigo = len(lista_codigos) + 1  # lendo o codigo + 1
    codigo_str = "COD-{}".format(codigo)  # formatando o codigo
    lista_codigos.append((codigo_str, nome, placa, cep, veiculo, placa, data_criacao))  # montando a lista em seguencia
    lista_produto_padrao.append((placa, data_criacao))

    texto = f'''
            
    Pessoa Cadastrado : {nome}
    Tipo :  {veiculo}
    Placa Veiculo : {placa}
    Numero do Chassi: {chassi}
    codigo gerado: {codigo} 

    '''
    texto_cadastro["text"] = texto



#______________________________ENTRADA DE DADOS BOTOES E TEXTOS________________________________________________


#  1 ENTRADA DE NOME:
janela_nome = Label(text=" NOME: ")
janela_nome.grid(row=1, column=0, padx=0, pady=0, sticky='nswe')
entrada = Entry()
entrada.grid(row=1, column=1, pady=0, padx=0, sticky='nswe')

# 2 ENTRADA  CAMPO CEP :
janela_cep = Label(text=" Digite seu CEP:  ")
janela_cep.grid(row=2, column=0, padx=0, pady=0, sticky='nswe')
entrada_cep = Entry()
entrada_cep.grid(row=2, column=1, pady=0, padx=0, sticky='nswe')

botao_codigo = tkinter.Button(text="CONSULTAR CEP. ", command=cep_busca)
botao_codigo.grid(row=3, column=0, pady=2, padx=5, sticky='nswe', columnspan=3)

# 3 ENTRADA TIPO DE VEICULO:

janela_veiculo = Label(text="TIPO DE VEICULO:  ")
janela_veiculo.grid(row=4, column=0, padx=0, pady=0, sticky='nswe')

botao_combobox = ttk.Combobox(values=lista_tipos)
botao_combobox.grid(row=4, column=1, pady=0, padx=0, sticky='nswe')

# 4 ENTRADA PLACA DE VEICULO

placa = Label(text='PLACA DO VEICULO: ')
placa.grid(row=1, column=3, pady=0, padx=0, sticky='nswe')
placa_veiculo = Entry()
placa_veiculo.grid(row=1, column=4, pady=0, padx=10, sticky='nswe')

consulta_placa = Label(text="DESEJA CONSULTA DETRAN: ")
consulta_placa.grid(row=2, column=3, pady=0, padx=0, sticky='nswe')
botao_pag_detran = tkinter.Button(text=" CONSULTAR ", command=funcao_chamar_pag_consulta_veiculo.janela_api)
botao_pag_detran.grid(row=2, column=4, pady=5, padx=5, sticky='nswe')

# CHASSI DO VEICULO

janela_chassi = Label(text='NUMERO DO CHASSI:')
janela_chassi.grid(row=6, column=0, padx=0, pady=0, sticky='nswe')
entrada_chassi = Entry()
entrada_chassi.grid(row=6, column=1, pady=0, padx=0, sticky='nswe')

# GERA CODIGO

botao_codigo = tkinter.Button(text="Click para gerar codigo. ", command=inserir_codigo)
botao_codigo.grid(row=9, column=0, pady=2, padx=5, sticky='nswe', columnspan=3)
texto_cadastro = Label(janela, text=" ")  # MENSAGEM EM TEXTO , CHAMANDO A FUNCAO .
texto_cadastro.grid(row=10, column=1, sticky='nswe')

#______________________________MESAGEM TELA DE SERVICO _CADASTRADO________________________________________
def get_dados():
    nome = entrada.get()
    cep = entrada_cep.get()
    placa = placa_veiculo.get()
    chassi = entrada_chassi.get()
    t_veiculo = botao_combobox.get()

    banco_de_dados.append((nome, cep, placa, chassi, t_veiculo))
    print(banco_de_dados)
    messagebox.showinfo('SERVIÇO CADASTRADO.')


# BOTAO CONFIRMAR CASATRO DE TUDO

botao_codigo = tkinter.Button(text=" Confirmar ", command=get_dados)
botao_codigo.grid(row=11, column=0, pady=5, padx=5, sticky='nswe')

# BOTAO CANCELAR

botao_codigo = tkinter.Button(text=" Cancelar ")
botao_codigo.grid(row=11, column=1, pady=5, padx=5, sticky='nswe')

#_________________________________JANELA 2 _______________________________________________________________

def abrir_jabela():
    # janela_2 = tkinter.Toplevel()
    # janela_2.title('nova janela')

    janela2 = Tk()
    janela2.geometry("800x800")
    janela2.title('DESPACHANTE DIGITAL - CHECK LIST ')
    janela2.resizable = True  # resizable=True movintar a tela e false trava tela
    janela2.confirm_close = True  # confirm_close=True, pergunta de confirmaçao fechar janela

    # --TEXTO TELA----
    texto = Label(text="PREENCHA O CHECK LIST")
    texto.grid(row=0, column=1, pady=0, padx=0, sticky='nswe')

    #  1 ENTRADA DE NOME:
    entrada = Label(janela2, text=" NOME: ")
    entrada.grid(row=3, column=0, padx=0, pady=0, sticky='nswe')
    entrada1 = Entry(janela2)
    entrada1.grid(row=3, column=1, pady=0, padx=0, sticky='nswe')

    # 2 ENTRADA PLACA:
    entrada = Label(janela2, text=" DIGITE A PLACA :  ")
    entrada.grid(row=4, column=0, padx=0, pady=0, sticky='nswe')
    placa = Entry(janela2)
    placa.grid(row=4, column=1, pady=0, padx=0, sticky='nswe')

    def bd_tela_2():
        banco = dict()
        dados1 = list()

        banco['id'] = entrada1.get()
        banco['placa'] = placa.get()

        dados1.append(banco.copy())
        print(f' ESTE E O PRINTE list()  {dados1}')
        print(f' ESTE E O PRINTE dict()  {banco}')

    botao_confirma = tkinter.Button(janela2, text=" CONFIRMA_2 ", command=bd_tela_2)
    botao_confirma.grid(row=5, column=0, pady=5, padx=5, sticky='nswe')

    botao_cancelar = tkinter.Button(janela2, text=" cancelar_2 ", command=janela2.destroy)
    botao_cancelar.grid(row=5, column=1, pady=5, padx=5, sticky='nswe')


                # _____________________CHAMAR JANELA 2________________________


botao_codigo = tkinter.Button(text="Janela 2", command=abrir_jabela)
botao_codigo.grid(row=12, column=1, pady=5, padx=5, sticky='nswe')

botao_sair = tkinter.Button(text=" SAIR ")
botao_sair.grid(row=8, column=1, pady=5, padx=5, sticky='nswe')


# ________________________________ JANELA 3 ________________________________________________________


def abrir_jabela_3():


    janela3 = Tk()
    janela3.geometry("800x800")
    janela3.title('DESPACHANTE DIGITAL - CHECK LIST ')
    janela3.resizable = True  # resizable=True movintar a tela e false trava tela
    janela3.confirm_close = True  # confirm_close=True, pergunta de confirmaçao fechar janela



    nome = StringVar()
    fone = StringVar()
    sexo= StringVar()
    idade = StringVar()
    endereco =StringVar()


    def clear ():
        nome.delete(0,END)
        fone.delete(0,END)
        idade.delete(0,END)
        endereco.delete(0,END)


     #def pode_pular()  #---> funsao deixar passar
        #pass


# ________ ESTE E O CODIGO PARA CRIAR UM BANCO DE DADOS_______________________________
    documentos = pathlib.Path ('Banco de dados.xlsx')
    if documentos.exists():
        pass
    else:
        documentos= Workbook()
        sheet = documentos.active
        sheet['A1'] = "nome"
        sheet['B1'] = "fone"
        sheet['C1'] = "sexo"
        sheet['D1'] = "idade"
        sheet['E1'] = "endereco"

        documentos.save('Banco de dados.xlsx')

    def puxando_dados():

        nomeGet = nome.get()
        foneGet = fone.get()
        sexoGet = sexo.get()
        idadeGet = idade.get()
        enderecoGet= endereco.get()

        documentos=openpyxl.load_workbook(r'Banco de dados.xlsx')
        sheet =documentos.active

        sheet.cell(column= 1, row=sheet.max_row + 1, value=nomeGet)
        sheet.cell(column= 2, row=sheet.max_row , value=foneGet)
        sheet.cell(column= 3, row=sheet.max_row , value=idadeGet)
        sheet.cell(column= 4, row=sheet.max_row , value=sexoGet)
        sheet.cell(column= 5, row=sheet.max_row ,value= enderecoGet)

        documentos.save('Banco de dados.xlsx')



        print(nomeGet,foneGet,idadeGet,sexoGet,enderecoGet)




        # --TEXTO TELA----
    texto = Label(janela3,text="DADOS A SER PREENCHIDOS ")
    texto.grid(row=0, column=1, pady=0, padx=0 ,sticky='nswe')

        #  ENTRADA DE NOME:
    nome = Label(janela3,text=" NOME: ")
    nome.grid(row=2, column=0, padx=0, pady=0, sticky='nswe')
    nome = Entry(janela3)
    nome.grid(row=2, column=1, pady=0, padx=0, sticky='nswe')


         #  ENTRADA DE Telefone:
    fone = Label(janela3,text=" TELEFONE: ")
    fone.grid(row=3, column=0, padx=0, pady=0, sticky='nswe')
    fone = Entry(janela3)
    fone.grid(row=3, column=1, pady=0, padx=0, sticky='nswe')


         #  ENTRADA DE GENERO IDADE:
    idade = Label(janela3,text=" IDADE: ")
    idade.grid(row=4, column=0, padx=0, pady=0, sticky='nswe')
    idade = Entry(janela3)
    idade.grid(row=4, column=1, pady=0, padx=0,sticky='nswe')


         # ENTRADA DE SEXO:
    sexo = Label(janela3,text=" SEXO : ")
    sexo.grid(row=4, column=2, padx=0, pady=0, sticky='nswe')

    sexo= ttk.Combobox(janela3,values=['Masculino','Femenino'])
    sexo.grid(row=4, column=3, pady=0, padx=0, sticky= 'nswe')
    sexo.set('Masculino')



        #  ENTRADA ENDERECO:
    endereco = Label(janela3, text=" ENDERECO:  ")
    endereco.grid(row=6, column=0, padx=0, pady=0, sticky='nswe')
    endereco = Entry(janela3)
    endereco.grid(row=6 , column=1 , pady=0, padx=0,ipady= 20, sticky='nswe',)


    # BOTOES DE CONFIRMACAO

    botao_codigo = tkinter.Button(janela3,text="SALVAR",command=puxando_dados)
    botao_codigo.grid(row=7, column=1, pady=5, padx=5, sticky='nswe')

    botao_codigo = tkinter.Button(janela3,text="LIMPAR",command=clear)
    botao_codigo.grid(row=7, column=2, pady=5, padx=5, sticky='nswe')

    botao_codigo = tkinter.Button(janela3,text="SAIR", command=janela3.destroy)
    botao_codigo.grid(row=7, column=3, pady=5, padx=5, sticky='nswe')









botao_codigo = tkinter.Button(text="Janela 3", command=abrir_jabela_3)
botao_codigo.grid(row=14, column=0, pady=5, padx=5, sticky='nswe')

botao_sair = tkinter.Button(text=" SAIR ")
botao_sair.grid(row=14, column=1, pady=5, padx=5, sticky='nswe')


#________________________________FIM DO LOOP DA JANELA PRINCIPAL ____________________________________

janela.mainloop()

#_________________________________FIM EXECUCAO do Programa ___________________________________________


#________________________________ IMPRESSAO PARA CONTROLE DE FUNCIONAMENTO _BACK AND__________
for x in lista_codigos:
    print(f'casadastro =  {x}')

# botao_combobox1 =  ttk.Menubutton(menu = ['celso','cleia','neide'])
# botao_combobox1.grid(row=5, column=0, pady=0, padx=0, sticky= 'nswe', columnspan=1)()
