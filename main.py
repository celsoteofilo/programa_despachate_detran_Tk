# _____________ Bibiliotecas para FRONT___________________________________________
import pathlib
import tkinter
from tkinter import *  # importa janelas
from tkinter import ttk, messagebox # importa o cromobox -> rolagem de dados

# _______________________Biblioteca para BANCO DE DADOS___________________________
from tkinter.ttk import Combobox
import openpyxl, xlrd
from openpyxl import Workbook
import pathy

# ______________________BIBLIOTECA  para DATA e HORA para o sistema _______________
import datetime as dt

# ___________________BIBLIOTECA PARA FUNCOES INTERNAS _____________________________
from tqdm import tk

import funcao_chama_habilitacao
import funcao_chamar_pag_consulta_veiculo

#cores ------------------------------
co0 = "#fof3f5" # preto
co1 = "#feffff" # branco
co2 = "#3fb5a3" # verde
co3 = "#38576b" # valor
co4 = "#403d3d" # letra
# ____________________LISTAS GERADAS PRINCIPAIS PARA CONSUMO  DO SISTEMA ___________

lista_codigos = []
lista_produto_padrao = []
banco_de_dados = []

# _____________________JANELA PRINCIPAL ._____________________________________________

janela = Tk()
janela.geometry("920x800+1000+300") # tamnanho da dela + centralizaçao da tela
janela.title('DESPACHANTE DIGITAL')
janela.resizable = True  # resizable=True movintar a tela e false trava tela
janela.confirm_close = True  # confirm_close=True, pergunta de confirmaçao fechar janela

# ________________________________CRIANDO A IMAGEM ________

bg = PhotoImage(file="v.png")
label1 = Label(janela, image=bg)
label1.place(x=250, y=500)

# ________________________________ JANELA CADASTRO __________________________________________________
def abrir_jabela_3():


    janela3 = Tk()
    janela3.geometry("900x800+1000+300")
    janela3.title('DESPACHANTE DIGITAL - CADASTRO ')
    janela3.resizable = True  # resizable=True movintar a tela e false trava tela
    janela3.confirm_close = True  # confirm_close=True, pergunta de confirmaçao fechar janela


    nome = StringVar()
    fone = StringVar()
    entrada_cep =  StringVar()
    cpf = StringVar()
    endereco = StringVar()
    numero = StringVar()
    cidade = StringVar()
    bairro = StringVar()
    habilitacao = StringVar()
    estado = StringVar()
    tipo_combustivel = StringVar()
    placa_veiculo = StringVar()
    chassi = StringVar()
    caixa_dialogo_a = StringVar()
    potencia = StringVar()
    veiculo_a= StringVar()

    def clear ():


        nome.delete(0,END)
        fone.delete(0,END)
        entrada_cep.delete(0,END)
        cpf.delete(0,END)
        endereco.delete(0,END)
        numero.delete(0,END)
        cidade.delete(0,END)
        bairro.delete(0,END)
        habilitacao.delete(0,END)
        estado.delete(0,END)
        tipo_combustivel.delete(0,END)
        placa_veiculo.delete(0,END)
        chassi.delete(0,END)
        caixa_dialogo_a.delete(0,END)
        potencia.delete(0,END)
        veiculo_a.delete(0,END)

     #def pode_pular()  #---> funsao deixar passar
        #pass

# ________ ESTE E O CODIGO PARA CRIAR UM BANCO DE DADOS_______________________________
    documentos = pathlib.Path ('Banco de dados.xlsx')
    if documentos.exists():
        pass
    else:
        documentos= Workbook()
        sheet = documentos.active

        sheet['A1'] ="nome"
        sheet['B1'] ="fone"
        sheet['C1'] ="etrada_cep"
        sheet['D1'] ="cpf"
        sheet['E1'] ="endereco"
        sheet['F1'] ="numero"
        sheet['G1'] ="cidade"
        sheet['H1'] ="bairro"
        sheet['I1'] ="habilitacao"
        sheet['J1'] ="estado"
        sheet['K1'] ="tipo_combustivel"
        sheet['L1'] ="placa_veiculo"
        sheet['M1'] ="chassi"
        sheet['N1'] ="caixa_dialogo_a"
        sheet['O1'] ="potencia"
        sheet['P1'] ="veiculo_a"


        documentos.save('Banco de dados.xlsx')

    def puxando_dados():

        nomeGet = nome.get()
        foneGet = fone.get()
        cepGet = entrada_cep.get()
        cpfGet = cpf.get()
        enderecoGet =endereco.get()
        numeroGet =numero.get()
        cidadeGet =cidade.get()
        bairroGet =bairro.get()
        habilitacaoGet = habilitacao.get()
        estadoGet =estado.get()
        tipo_combustivelGet =tipo_combustivel.get()
        placa_veiculoGet =placa_veiculo.get()
        chassiGet =chassi.get()
        caixa_dialogoGet =caixa_dialogo_a.get()
        potenciaGet=potencia.get()
        veiculoGet=veiculo_a.get()

        documentos=openpyxl.load_workbook(r'Banco de dados.xlsx')
        sheet =documentos.active

        sheet.cell(column= 1, row=sheet.max_row + 1, value=nomeGet)
        sheet.cell(column= 2, row=sheet.max_row , value=foneGet)
        sheet.cell(column= 3, row=sheet.max_row , value=cepGet)
        sheet.cell(column= 4, row=sheet.max_row , value=cpfGet)
        sheet.cell(column= 5, row=sheet.max_row ,value= enderecoGet)
        sheet.cell(column= 6, row=sheet.max_row , value=numeroGet)
        sheet.cell(column= 7, row=sheet.max_row , value=cidadeGet)
        sheet.cell(column= 8, row=sheet.max_row , value=bairroGet)
        sheet.cell(column= 9, row=sheet.max_row , value=habilitacaoGet)
        sheet.cell(column= 10, row=sheet.max_row , value=estadoGet)
        sheet.cell(column= 11, row=sheet.max_row , value=tipo_combustivelGet)
        sheet.cell(column= 12, row=sheet.max_row , value=placa_veiculoGet)
        sheet.cell(column= 13, row=sheet.max_row , value=chassiGet)
        sheet.cell(column= 14, row=sheet.max_row , value=caixa_dialogoGet)
        sheet.cell(column= 15, row=sheet.max_row , value=potenciaGet)
        sheet.cell(column= 16, row=sheet.max_row , value=veiculoGet)


        documentos.save('Banco de dados.xlsx')

        print(nomeGet,foneGet,cpfGet,numeroGet,cidadeGet,bairroGet,habilitacaoGet,estadoGet,tipo_combustivelGet,
              placa_veiculoGet,chassiGet,potenciaGet,veiculoGet)
 #_________________CONTRUCAO DE CODIGO DE CADASTRO____________N OS_____________________________


    def ordem_servico():

            # __________________puxando data e hora_____________
        data_criacao = dt.datetime.now()
        data_criacao.strftime('%y/%m/%d--%H:%M')

            #____________________ criando um codigo_____________
        codigo = len(lista_codigos) + 1  # lendo o codigo + 1
        codigo_str = "COD-{}".format(codigo)  # formatando o codigo
        lista_codigos.append((codigo_str, data_criacao))  # montando a lista em seguencia
        lista_produto_padrao.append((data_criacao))

        texto = f'''
                    
        ORDEM DE SERVICO: {codigo} 
        DATA : {data_criacao}
        
        '''
        texto_OS["text"] = texto

    texto_OS = Label(janela3, text=" ")
    texto_OS.place(x=450 , y=400)
#----------------------------------------------------------------------------------------------------------------------
        # --TEXTO TELA----
    texto = Label(janela3,text="DADOS A SER PREENCHIDOS ",font='arial 15 bold')
    texto.grid(row=0, column=1, pady=0, padx=0 ,sticky='nswe')

        # LINHA VERDE 0

    linha0 = Label (janela3, text=" ",width=350,anchor=N,font=('ily 1'), bg = co2)
    linha0.place(x=460,y=7)


        #  NOME:
    nome = Label(janela3,text=" NOME: ")
    nome.grid(row=2, column=0, padx=0, pady=0, sticky='nswe')
    nome = Entry(janela3)
    nome.grid(row=2, column=1, pady=0, padx=0, sticky='nswe')


         #  Telefone:
    fone = Label(janela3,text=" TELEFONE: ")
    fone.grid(row=2, column=2, padx=0, pady=0, sticky='nswe')
    fone = Entry(janela3)
    fone.grid(row=2, column=3, pady=0, padx=0, sticky='nswe')


    # 2 ENTRADA  CAMPO CEP :

    def cep_busca_3():

        import requests

        cep = entrada_cep.get()
        cep = cep.replace("-", "").replace(".", "").replace(" ", "")

        if len(cep) == 8:
            link = f'https://viacep.com.br/ws/{cep}/json/'
            requisicao = requests.get(link)
            print(requisicao)

            resposta = requisicao.json()

            uf = resposta['uf']
            cidade = resposta['localidade']
            bairro = resposta['bairro']
            logradouro = resposta['logradouro']
            complemento = resposta['complemento']

            print(resposta)
        else:
            messagebox.showinfo('<CEP>, INVALIDO!!!')

        texto_cep = f'''
       
        CEP INFORMADO : {cep}
        RUA: {logradouro}
        BAIRRO : {bairro}
        CIDADE :  {cidade}
        COMPLEMENTO {complemento}
        UF : {uf}
       
        '''
        texto_cadastro["text"] = texto_cep

    cep = Label(janela3,text=" Digite seu CEP:  ")
    cep.grid(row=4, column=0, padx=0, pady=0, sticky='nswe')
    entrada_cep = Entry(janela3)
    entrada_cep.grid(row=4, column=1, pady=0, padx=0, sticky='nswe')

    botao_codigo = tkinter.Button(janela3,text="CONSULTAR CEP. ", command=cep_busca_3)
    botao_codigo.grid(row=5, column=2, pady=0, padx=0, sticky='nswe', columnspan=1)


        # ___ TEXTO DO CePNa TeLA ___________


    texto_cadastro = Label(janela3, text=" ")
    texto_cadastro.place(x=170 , y= 360)


        # __________________CPF:________________________


    cpf = Label(janela3, text="CPF:")
    cpf.grid(row=4, column=2, padx=0, pady=0, sticky='nswe')
    cpf = Entry(janela3)
    cpf.grid(row=4 , column=3 , pady=0, padx=0,ipady= 0, sticky='nswe')

        # _________________________ ENDERECO:__________________
    endereco = Label(janela3, text="ENDERECO: ")
    endereco.grid(row=6, column=0, padx=0, pady=0, sticky='nswe')
    endereco = Entry(janela3)
    endereco.grid(row=6 , column=1 , pady=0, padx=0,ipady= 0, sticky='nswe')

    #____________  NUMERO DA CASA:_______________________________

    numero = Label(janela3, text=" Nº:  ")
    numero.grid(row=6, column=2, padx=0, pady=0, sticky='nswe')
    numero = Entry(janela3)
    numero.grid(row=6 , column=3 , pady=0, padx=0,ipady= 0,sticky='nswe')

    # __________________________  CIDADE :_____________________________________
    lista_cidades =sorted(['BELO HORIZONTE ','IBIRITE','SARZEDO','MARIO CAMPOS','BETIM','S.J DE BICAS','IGARAPE','BRUMADINHO',
                    'CRUCILANDIA','PATOS DE MINAS', 'JANUARIA','CONGONHAS','CORINTOS','SETE LAGOS',
                    'SABARA','PIEDADE GERAIS'])

    cidade = Label(janela3,text="CIDADE:  ")
    cidade.grid(row=7, column=0, padx=0, pady=0, sticky='nswe')
    cidade= ttk.Combobox(janela3,values=lista_cidades)
    cidade.grid(row=7, column=1, pady=0, padx=0, sticky='nswe')

    # ___________________________BAIRRO___________________________________________

    bairro = Label(janela3, text=" BAIRROº:  ")
    bairro.grid(row=7, column=2, padx=0, pady=0, sticky='nswe')
    bairro = Entry(janela3)
    bairro.grid(row=7 , column=3 , pady=0, padx=0,ipady= 0,sticky='nswe')



    # ___________________________ ENTRADA HABILTACAO :___________-------------
    habilitacao = Label(janela3,text=" Nº HABILITACAO: ")
    habilitacao.grid(row=9, column=0, padx=0, pady=0, sticky='nswe')
    habilitacao = Entry(janela3)
    habilitacao.grid(row=9, column=1, pady=0, padx=0,sticky='nswe')


    # ____________________________ENTRADA DE ESTADO:__________________________________
    estado = Label(janela3,text=" ESTADO UF: ")
    estado.grid(row=9, column=2, padx=0, pady=0, sticky='nswe')

    estado= ttk.Combobox(janela3,values=sorted(['MG','SP','RJ','PE','SE','DF','TO','MA','MT','SC','RS','PA']))
    estado.grid(row=9, column=3, pady=0, padx=0, sticky= 'nswe')
    estado.set('MG')


    #______________________ BOTOES DE CONFIRMACAO  CADASTRO- JANELA 3 _______________________________


    botao_salva = tkinter.Button(janela3,text="SALVAR",command= puxando_dados and ordem_servico)
    botao_salva.place(x=100, y=700, width=200)




    botao_limpa = tkinter.Button(janela3,text="LIMPAR",command=clear)
    botao_limpa.place(x=300, y=700, width=200)

    botao_sair = tkinter.Button(janela3,text="SAIR", command=janela3.destroy)
    botao_sair.place(x=500, y=700, width=200)


    #Linha1 = Label (janela3, text="",width=275,anchor=NW,font=('ily 1'), bg = co2)
    #Linha1.place(x=10,y=45)


    Linha1 = Label (janela3, text=" ",width=800,anchor=N,font=('ily 1'), bg = co2)
    Linha1.place(x=0,y=240)



    #____________________________JANELA 3 CADASTRO DE CARRO _______________________________


    # ENTRADA DE  TIPO DE CARRO
    texto = Label(janela3,text="CADASTRO DE CARRO ",font='arial 15 bold')
    texto.grid(row=16, column=1, pady=0, padx=0 ,sticky='nswe')


    veiculo = Label(janela3,text="TIPO DE VEICULO:  ")
    veiculo.grid(row=18, column=0, padx=0, pady=0, sticky='nswe')

    veiculo_a= ttk.Combobox(janela3,values=sorted(["CARRO", "MOTO", "CAMINHAO ","OUTROS"]))
    veiculo_a.grid(row=18, column=1, pady=0, padx=0, sticky='nswe')


    # ENTRADA DE TIPO DE COMBUSTIVEL ______________________________________________________

    tipo_combustivel = Label(janela3,text="TIPO COMBUSTIVEL:  ")
    tipo_combustivel.grid(row=18, column=2, padx=0, pady=0, sticky='nswe')
    tipo_combustivel = ttk.Combobox(janela3,values=sorted(["GASOLINA ", "ALCOOL", "DISEL","ELETRICO","GAS","FLEX",
                                                           "FLEX + GAS","GASOLINA +  GAS", "ALCOOL + GAS"]))
    tipo_combustivel.grid(row=18, column=3, pady=0, padx=0, sticky='nswe')

    # 4 ENTRADA PLACA DE VEICULO

    placa = Label(janela3,text='PLACA DO VEICULO: ')
    placa.grid(row=19, column=0, pady=0, padx=0, sticky='nswe')
    placa_veiculo = Entry(janela3)
    placa_veiculo.grid(row=19, column=1, pady=0, padx=0, sticky='nswe')

      # CHASSI DO VEICULO

    chassi = Label(janela3,text='NUMERO DO CHASSI:')
    chassi.grid(row=19, column=2, padx=0, pady=0, sticky='nswe')
    chassi = Entry(janela3)
    chassi.grid(row=19, column=3, pady=0, padx=0, sticky='nswe')

    # POTENCIA MOTOR

    potencia = Label(janela3,text="POT. MOTOR:  ")
    potencia.grid(row=21, column=2, padx=0, pady=0, sticky='nswe')
    potencia = ttk.Combobox(janela3,values=["1.0 ", "1.4","1.5","1.8","2.0","2.8","3.0","4.0","5.0","OUTROS"])
    potencia.grid(row=21, column=3, pady=0, padx=0, sticky='nswe')

    # PLACAS

    checar = Label(janela3,text=' PLACA:')
    checar.grid(row=19, column=0 , sticky='nswe')

    # CAIXA DE DIALOGO

    caixa_dialogo = Label(janela3,text='OBS NO CADASTRO : ',font='arial 15 bold',)
    caixa_dialogo.place(x=200 , y=550, )
    caixa_dialogo_a =Entry(janela3 )
    caixa_dialogo_a.place(x= 450, y= 550,width=300, height=100)

    # CONSULTA DETRAN

    consulta_placa = Label(janela3,text="CONSULTA DETRAN VEICULO: ")
    consulta_placa.grid(row=21, column=0, pady=0, padx=0, sticky='nswe')
    botao_pag_detran = tkinter.Button(janela3,text=" CONSULTAR ", command=funcao_chamar_pag_consulta_veiculo.janela_api)
    botao_pag_detran.grid(row=21, column=1, pady=5, padx=5, sticky='nswe')

    # CONSULTA HABILITACAO

    consulta_habilitacao = Label(janela3,text="CONSULTA DETRAN HABILITACAO: ")
    consulta_habilitacao.grid(row=22, column=0, pady=0, padx=0, sticky='nswe')
    botao_detran = tkinter.Button(janela3,text=" CONSULTAR ", command=funcao_chama_habilitacao.janela_habilitacao)
    botao_detran.grid(row=22, column=1, pady=5, padx=5, sticky='nswe')


#_________________________________JANELA SERVICOS _____________________________________________

def abrir_janela_4():

    janela4 = Tk()
    janela4.geometry("800x800+1000+300")
    janela4.title('DESPACHANTE DIGITAL - CHECK LIST ')
    janela4.resizable = True  # resizable=True movintar a tela e false trava tela
    janela4.confirm_close = True  # confirm_close=True, pergunta de confirmaçao fechar janela

        # --TEXTO TELA----
    texto = Label(janela4,text="PREENCHA O CHECK LIST")
    texto.grid(row=0, column=1, pady=0, padx=0, sticky='nswe')

        #  1 ENTRADA DE NOME:
    entrada = Label(janela4, text=" NOME: ")
    entrada.grid(row=3, column=0, padx=0, pady=0, sticky='nswe')
    entrada = Entry(janela4)
    entrada.grid(row=3, column=1, pady=0, padx=0, sticky='nswe')

        # 2 ENTRADA PLACA:
    placa = Label(janela4, text=" DIGITE A PLACA :  ")
    placa.grid(row=4, column=0, padx=0, pady=0, sticky='nswe')
    placa = Entry(janela4)
    placa.grid(row=4, column=1, pady=0, padx=0, sticky='nswe')

    def bd_tela_4():
        banco = dict()
        dados1 = list()

        banco['id'] = entrada.get()
        banco['placa'] = placa.get()

        dados1.append(banco.copy())
        print(f' ESTE E O PRINTE list()  {dados1}')
        print(f' ESTE E O PRINTE dict()  {banco}')

    botao_confirma = tkinter.Button(janela4, text=" CONFIRMA_2 ", command=bd_tela_4)
    botao_confirma.grid(row=5, column=0, pady=5, padx=5, sticky='nswe')

    botao_cancelar = tkinter.Button(janela4, text=" cancelar_2 ", command=janela4.destroy)
    botao_cancelar.grid(row=5, column=1, pady=5, padx=5, sticky='nswe')


#_________________________________JANELA 5 ___GERENCIAMENTO ___________________________________________
def abrir_janela_5():

    janela5 = Tk()
    janela5.geometry("800x800+1000+300")
    janela5.title('DESPACHANTE DIGITAL - CHECK LIST ')
    janela5.resizable = True  # resizable=True movintar a tela e false trava tela
    janela5.confirm_close = True  # confirm_close=True, pergunta de confirmaçao fechar janela

        # --TEXTO TELA----
    texto = Label(janela5,text="PREENCHA O CHECK LIST")
    texto.grid(row=0, column=1, pady=0, padx=0, sticky='nswe')

        #  1 ENTRADA DE NOME:
    entrada = Label(janela5, text=" NOME: ")
    entrada.grid(row=3, column=0, padx=0, pady=0, sticky='nswe')
    entrada = Entry(janela5)
    entrada.grid(row=3, column=1, pady=0, padx=0, sticky='nswe')

        # 2 ENTRADA PLACA:
    placa = Label(janela5, text=" DIGITE A PLACA :  ")
    placa.grid(row=4, column=0, padx=0, pady=0, sticky='nswe')
    placa = Entry(janela5)
    placa.grid(row=4, column=1, pady=0, padx=0, sticky='nswe')

    def bd_tela_5():
        banco = dict()
        dados1 = list()

        banco['id'] = entrada.get()
        banco['placa'] = placa.get()

        dados1.append(banco.copy())
        print(f' ESTE E O PRINTE list()  {dados1}')
        print(f' ESTE E O PRINTE dict()  {banco}')

    botao_confirma = tkinter.Button(janela5, text=" CONFIRMA_2 ", command=bd_tela_5)
    botao_confirma.grid(row=5, column=0, pady=5, padx=5, sticky='nswe')

    botao_cancelar = tkinter.Button(janela5, text=" cancelar_2 ", command=janela5.destroy)
    botao_cancelar.grid(row=5, column=1, pady=5, padx=5, sticky='nswe')


#____________________________________JANELA __6_ FINANCEIRO_______________________________________

def abrir_janela_6():
    janela6 = Tk()
    janela6.geometry("800x800+1000+300")
    janela6.title('DESPACHANTE DIGITAL - CHECK LIST ')
    janela6.resizable = True  # resizable=True movintar a tela e false trava tela
    janela6.confirm_close = True  # confirm_close=True, pergunta de confirmaçao fechar janela

    # --TEXTO TELA----
    texto = Label(janela6, text="PREENCHA O CHECK LIST")
    texto.grid(row=0, column=1, pady=0, padx=0, sticky='nswe')

    #  1 ENTRADA DE NOME:
    entrada = Label(janela6, text=" NOME: ")
    entrada.grid(row=3, column=0, padx=0, pady=0, sticky='nswe')
    entrada = Entry(janela6)
    entrada.grid(row=3, column=1, pady=0, padx=0, sticky='nswe')

    # 2 ENTRADA PLACA:
    placa = Label(janela6, text=" DIGITE A PLACA :  ")
    placa.grid(row=4, column=0, padx=0, pady=0, sticky='nswe')
    placa = Entry(janela6)
    placa.grid(row=4, column=1, pady=0, padx=0, sticky='nswe')

    def bd_tela_6():
        banco = dict()
        dados1 = list()

        banco['id'] = entrada.get()
        banco['placa'] = placa.get()

        dados1.append(banco.copy())
        print(f' ESTE E O PRINTE list()  {dados1}')
        print(f' ESTE E O PRINTE dict()  {banco}')

    botao_confirma = tkinter.Button(janela6, text=" CONFIRMA_2 ", command=bd_tela_6)
    botao_confirma.grid(row=5, column=0, pady=5, padx=5, sticky='nswe')

    botao_cancelar = tkinter.Button(janela6, text=" cancelar_2 ", command=janela6.destroy)
    botao_cancelar.grid(row=5, column=1, pady=5, padx=5, sticky='nswe')

#_____________________________________JANELA 7__MANUTECAO______________________________________

def abrir_janela_7():
    janela7 = Tk()
    janela7.geometry("800x800+1000+300")
    janela7.title('DESPACHANTE DIGITAL - CHECK LIST ')
    janela7.resizable = True  # resizable=True movintar a tela e false trava tela
    janela7.confirm_close = True  # confirm_close=True, pergunta de confirmaçao fechar janela

    # --TEXTO TELA----
    texto = Label(janela7, text="PREENCHA O CHECK LIST")
    texto.grid(row=0, column=1, pady=0, padx=0, sticky='nswe')

    #  1 ENTRADA DE NOME:
    entrada = Label(janela7, text=" NOME: ")
    entrada.grid(row=3, column=0, padx=0, pady=0, sticky='nswe')
    entrada = Entry(janela7)
    entrada.grid(row=3, column=1, pady=0, padx=0, sticky='nswe')

    # 2 ENTRADA PLACA:
    placa = Label(janela7, text=" DIGITE A PLACA :  ")
    placa.grid(row=4, column=0, padx=0, pady=0, sticky='nswe')
    placa = Entry(janela7)
    placa.grid(row=4, column=1, pady=0, padx=0, sticky='nswe')

    def bd_tela_7():
        banco = dict()
        dados1 = list()

        banco['id'] = entrada.get()
        banco['placa'] = placa.get()

        dados1.append(banco.copy())
        print(f' ESTE E O PRINTE list()  {dados1}')
        print(f' ESTE E O PRINTE dict()  {banco}')

    botao_confirma = tkinter.Button(janela7, text=" CONFIRMA_2 ", command=bd_tela_7)
    botao_confirma.grid(row=5, column=0, pady=5, padx=5, sticky='nswe')

    botao_cancelar = tkinter.Button(janela7, text=" cancelar_2 ", command=janela7.destroy)
    botao_cancelar.grid(row=5, column=1, pady=5, padx=5, sticky='nswe')


#____________________________________JANELA LOGIN___________________________________________




entrada_cred_nome = []
entrada_cred_senha = []

a_novo_nome =[]
a_nova_senha=[]

#________________FUNCAO  ENTRADA NOME E SENHA DE USUSARIO____________________

def botao_entrada():
    nome = entra_nome.get()
    senha = entra_senha.get()


    entrada_cred_nome.append(nome)
    entrada_cred_senha.append(senha)

    print(entrada_cred_nome, entrada_cred_senha)
    print('AS DUAS NOVAS ,ENTRADA ')

    if  entra_nome :
        print('SENHA VALIDA !!!')
        messagebox.showinfo('Seja Bem vindo !!!')

    else:
        print(False)
        print('CADASTRE SENHA INVALIDA !!!')
        messagebox.showinfo('NAO CADASTRADO!!! CADASTRE A BAIXO.')
        entrada_cred_senha.clear()
        entrada_cred_nome.clear()

#________________________ FUNCAO ENTRA DE NOVO NOME E  NOVA SENHA_______________________
def senha_nova():

    nome = novo_nome.get()
    senha = nova_senha.get()

    a_novo_nome.append(nome)
    a_nova_senha.append(senha)

    print(a_novo_nome,a_nova_senha)


# ___________________ENTRADA  NOME__________________________________

janela_nome= Label(text= "ENTRE COM NOME:  ")
janela_nome.place(x= 20, y = 150)
entra_nome = Entry()
entra_nome.place(x= 200, y = 150)

# _________________ENTRADA SENHA__________________________________

janela_senha= Label(text= "ENTRE COM SENHA  ")
janela_senha.place(x= 20, y = 180)
entra_senha = Entry()
entra_senha.place(x= 200, y = 180)

# __________________BOTAO CONFIRMAR ENTRADA__________________________

botao_confirmar_senha = tkinter.Button(text="ENTRAR ", command=botao_entrada)
botao_confirmar_senha.place(x= 300, y = 230)

#_______________________ENTRADAS SENHAS NOVAS ____________________________

#_________________________ENTRADA NOVO NOME_________________________________

janela_cadastro_nome = Label(text= " NOVO NOME ", fg="Red")
janela_cadastro_nome.place(x= 20, y = 270)
novo_nome= Entry()
novo_nome.place(x= 200, y = 270)


# __________________________ENTRADA NOVA SENHA_________________________________
janela_cadastro_senha = Label(text= " NOVA SENHA ",fg="Red")
janela_cadastro_senha.place(x= 20, y= 300)
nova_senha = Entry()
nova_senha.place(x= 200, y = 300)

# _______________________BOTAO CONFIRMA NOVA SENHA_______________________________________

botao_confirmar_senha = tkinter.Button(text="SALVAR. ",command=senha_nova)
botao_confirmar_senha.place(x= 300, y = 350)


#_______________________________________________BOTOES NA JANELA PRINCIPAL_____________________

botao_codigo = tkinter.Button(text="MANUTENCAO", command=abrir_janela_7)
botao_codigo.grid(row=1, column=4, pady=5, padx=5, sticky='nswe')

botao_codigo = tkinter.Button(text="FINACEIRO", command=abrir_janela_6)
botao_codigo.grid(row=1, column=3, pady=5, padx=5, sticky='nswe')

botao_codigo = tkinter.Button(text="GERENCIAMENTO", command=abrir_janela_5)
botao_codigo.grid(row=1, column=2, pady=5, padx=5, sticky='nswe')


botao_codigo = tkinter.Button(text="SERVICOS", command=abrir_janela_4)
botao_codigo.grid(row=1, column=1, pady=5, padx=5, sticky='nswe')

botao_codigo = tkinter.Button(text="CADASTRO", command=abrir_jabela_3)
botao_codigo.grid(row=1, column=0, pady=5, padx=5, sticky='nswe')

botao_sair = tkinter.Button(text=" SAIR ")
botao_sair.grid(row=1, column=5, pady=5, padx=5, sticky='nswe')


#________________________________FIM DO LOOP DA JANELA PRINCIPAL ____________________________________

janela.mainloop()

#_________________________________FIM EXECUCAO do Programa ___________________________________________


#________________________________ IMPRESSAO PARA CONTROLE DE FUNCIONAMENTO _BACK AND__________
for x in lista_codigos:
    print(f'casadastro =  {x}')

# botao_combobox1 =  ttk.Menubutton(menu = ['celso','cleia','neide'])
# botao_combobox1.grid(row=5, column=0, pady=0, padx=0, sticky= 'nswe', columnspan=1)()
